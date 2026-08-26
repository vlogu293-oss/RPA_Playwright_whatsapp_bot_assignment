import os
import re
import json
import time
from datetime import datetime

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# ============================================================
# CONFIGURATION
# ============================================================

EXCEL_FILE = "contacts.xlsx"

SCREENSHOT_DIR = "screenshots"
REPORT_DIR = "reports"

JSON_REPORT = os.path.join(REPORT_DIR, "whatsapp_report.json")
EXCEL_REPORT = os.path.join(REPORT_DIR, "whatsapp_report.xlsx")

# Persistent Chrome profile.
# This allows WhatsApp Web login to remain available
# after the first QR-code scan.
CHROME_PROFILE = os.path.abspath("whatsapp_profile")

DEFAULT_MESSAGE = "Hello {name}, this is a test message."

WAIT_AFTER_SEND = 3
WAIT_AFTER_CHAT_OPEN = 3


# ============================================================
# CREATE DIRECTORIES
# ============================================================

os.makedirs(SCREENSHOT_DIR, exist_ok=True)
os.makedirs(REPORT_DIR, exist_ok=True)
os.makedirs(CHROME_PROFILE, exist_ok=True)


# ============================================================
# UTILITY FUNCTIONS
# ============================================================

def clean_phone(phone):
    """
    Convert Excel phone value into a WhatsApp-friendly number.

    Example:
        +91 98765 43210
        -> 919876543210
    """

    if phone is None:
        return ""

    phone = str(phone).strip()

    # Keep digits only
    phone = re.sub(r"\D", "", phone)

    return phone


def safe_filename(text):
    """
    Make a safe filename.
    """

    text = str(text)

    return re.sub(r'[<>:"/\\|?*]', "_", text)


def timestamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ============================================================
# READ CONTACTS
# ============================================================

def read_contacts():

    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(
            f"Could not find {EXCEL_FILE}"
        )

    workbook = openpyxl.load_workbook(
        EXCEL_FILE,
        data_only=True
    )

    sheet = workbook.active

    headers = {}

    for cell in sheet[1]:
        if cell.value:
            headers[str(cell.value).strip().lower()] = cell.column

    required_columns = ["name", "phone", "message"]

    for column in required_columns:

        if column not in headers:

            raise ValueError(
                f"Missing required Excel column: {column}"
            )

    contacts = []

    for row in sheet.iter_rows(min_row=2, values_only=True):

        name = row[headers["name"] - 1]
        phone = row[headers["phone"] - 1]
        message = row[headers["message"] - 1]

        if not name or not phone:
            continue

        phone = clean_phone(phone)

        if not phone:
            continue

        if message:
            message = str(message)
        else:
            message = DEFAULT_MESSAGE

        message = message.replace(
            "{name}",
            str(name)
        )

        contacts.append({
            "name": str(name).strip(),
            "phone": phone,
            "message": message
        })

    return contacts


# ============================================================
# START CHROME
# ============================================================

def start_browser():

    options = Options()

    # Persistent profile
    options.add_argument(
        f"--user-data-dir={CHROME_PROFILE}"
    )

    options.add_argument("--start-maximized")

    # Prevent some unnecessary automation prompts
    options.add_argument(
        "--disable-blink-features=AutomationControlled"
    )

    driver = webdriver.Chrome(
        options=options
    )

    driver.get("https://web.whatsapp.com/")

    return driver


# ============================================================
# WAIT FOR WHATSAPP LOGIN
# ============================================================

def wait_for_login(driver):

    print()
    print("=" * 60)
    print("WhatsApp Web opened.")
    print()
    print("If a QR code is displayed:")
    print("1. Open WhatsApp on your phone.")
    print("2. Go to Linked Devices.")
    print("3. Scan the QR code.")
    print("=" * 60)
    print()

    while True:

        try:

            # Main WhatsApp application area
            driver.find_element(
                By.ID,
                "pane-side"
            )

            print("WhatsApp Web login detected.")
            break

        except Exception:
            print(
                "Waiting for WhatsApp login...",
                end="\r"
            )

            time.sleep(2)


# ============================================================
# OPEN CONTACT CHAT
# ============================================================

def open_chat(driver, phone):

    url = (
        "https://web.whatsapp.com/send"
        f"?phone={phone}"
        "&text="
    )

    driver.get(url)

    time.sleep(WAIT_AFTER_CHAT_OPEN)

    try:

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located(
                (
                    By.XPATH,
                    '//div[@contenteditable="true"]'
                )
            )
        )

        return True

    except Exception:

        print(
            f"Could not open chat for {phone}"
        )

        return False


# ============================================================
# FIND MESSAGE BOX
# ============================================================

def get_message_box(driver):

    possible_boxes = driver.find_elements(
        By.XPATH,
        '//div[@contenteditable="true"]'
    )

    if not possible_boxes:
        return None

    # Usually the last contenteditable is the message box
    return possible_boxes[-1]


# ============================================================
# SEND MESSAGE
# ============================================================

def send_message(driver, message):

    message_box = get_message_box(driver)

    if message_box is None:
        raise Exception(
            "WhatsApp message box was not found."
        )

    message_box.click()

    # Write message
    message_box.send_keys(message)

    time.sleep(1)

    # Send
    message_box.send_keys(Keys.ENTER)

    time.sleep(WAIT_AFTER_SEND)


# ============================================================
# SCREENSHOT
# ============================================================

def take_screenshot(driver, name, phone):

    filename = (
        f"{safe_filename(name)}_"
        f"{safe_filename(phone)}.png"
    )

    filepath = os.path.join(
        SCREENSHOT_DIR,
        filename
    )

    driver.save_screenshot(filepath)

    return filepath


# ============================================================
# EXTRACT LAST 3 INCOMING MESSAGES
# ============================================================

def extract_last_three_messages(driver):

    """
    Attempts to extract the latest 3 incoming messages.

    WhatsApp Web changes its DOM frequently, so several
    selectors are attempted.
    """

    messages = []

    selectors = [

        # Common WhatsApp message structure
        '//div[contains(@class,"message-in")]',

        # Alternative structure
        '//div[contains(@data-testid,"msg-container") and '
        'not(contains(@class,"message-out"))]',

        # Generic message containers
        '//div[contains(@class,"message-in")]'
        '//span[contains(@class,"selectable-text")]'
    ]

    elements = []

    for selector in selectors:

        try:

            elements = driver.find_elements(
                By.XPATH,
                selector
            )

            if elements:
                break

        except Exception:
            pass

    # Extract text
    for element in elements:

        try:

            text = element.text.strip()

            if not text:
                continue

            # Remove duplicates
            if text not in messages:
                messages.append(text)

        except Exception:
            continue

    # Return only last 3
    return messages[-3:]


# ============================================================
# VERIFY SENT MESSAGE
# ============================================================

def verify_message_sent(driver, message):

    """
    Basic verification that the sent text appears
    in the conversation.
    """

    try:

        xpath = (
            '//span[contains(@class,"selectable-text")]'
            f'[contains(normalize-space(.), '
            f'{json.dumps(message)})]'
        )

        elements = driver.find_elements(
            By.XPATH,
            xpath
        )

        return len(elements) > 0

    except Exception:

        return False


# ============================================================
# SAVE JSON REPORT
# ============================================================

def save_json_report(results):

    with open(
        JSON_REPORT,
        "w",
        encoding="utf-8"
    ) as file:

        json.dump(
            results,
            file,
            indent=4,
            ensure_ascii=False
        )


# ============================================================
# SAVE EXCEL REPORT
# ============================================================

def save_excel_report(results):

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "WhatsApp Report"

    headers = [
        "Name",
        "Phone",
        "Message Sent",
        "Status",
        "Timestamp",
        "Screenshot",
        "Last Message 1",
        "Last Message 2",
        "Last Message 3",
        "Error"
    ]

    sheet.append(headers)

    for result in results:

        last_messages = result.get(
            "last_3_messages",
            []
        )

        last_messages = (
            last_messages + ["", "", ""]
        )[:3]

        sheet.append([
            result.get("name", ""),
            result.get("phone", ""),
            result.get("message_sent", ""),
            result.get("status", ""),
            result.get("timestamp", ""),
            result.get("screenshot", ""),
            last_messages[0],
            last_messages[1],
            last_messages[2],
            result.get("error", "")
        ])

    # Adjust column widths
    for column in sheet.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:

                length = len(
                    str(cell.value)
                )

                if length > max_length:
                    max_length = length

            except Exception:
                pass

        sheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            60
        )

    workbook.save(EXCEL_REPORT)


# ============================================================
# PROCESS ONE CONTACT
# ============================================================

def process_contact(driver, contact):

    name = contact["name"]
    phone = contact["phone"]
    message = contact["message"]

    print()
    print("-" * 60)
    print(f"Processing: {name}")
    print(f"Phone: {phone}")
    print(f"Message: {message}")

    result = {
        "name": name,
        "phone": phone,
        "message_sent": message,
        "status": "failed",
        "timestamp": timestamp(),
        "screenshot": None,
        "last_3_messages": [],
        "error": None
    }

    try:

        # --------------------------------------------
        # Open WhatsApp chat
        # --------------------------------------------

        opened = open_chat(
            driver,
            phone
        )

        if not opened:

            raise Exception(
                "Unable to open WhatsApp chat."
            )

        # --------------------------------------------
        # Send message
        # --------------------------------------------

        send_message(
            driver,
            message
        )

        # --------------------------------------------
        # Verify message
        # --------------------------------------------

        sent = verify_message_sent(
            driver,
            message
        )

        if not sent:

            print(
                "Warning: message could not be verified."
            )

        else:

            print("Message sent successfully.")

        # --------------------------------------------
        # Screenshot
        # --------------------------------------------

        screenshot = take_screenshot(
            driver,
            name,
            phone
        )

        result["screenshot"] = screenshot

        # --------------------------------------------
        # Extract last 3 incoming messages
        # --------------------------------------------

        print(
            "Extracting last 3 messages..."
        )

        last_messages = (
            extract_last_three_messages(
                driver
            )
        )

        result["last_3_messages"] = last_messages

        # --------------------------------------------
        # Mark successful
        # --------------------------------------------

        result["status"] = "sent"

        print(
            f"Found {len(last_messages)} "
            f"incoming messages."
        )

    except Exception as error:

        result["error"] = str(error)

        print(
            f"ERROR: {error}"
        )

    result["timestamp"] = timestamp()

    return result


# ============================================================
# MAIN
# ============================================================

def main():

    print()
    print("=" * 60)
    print("WHATSAPP EXCEL AUTOMATION BOT")
    print("=" * 60)

    # --------------------------------------------
    # Read Excel
    # --------------------------------------------

    try:

        contacts = read_contacts()

    except Exception as error:

        print()
        print(
            f"Excel error: {error}"
        )

        return

    print()
    print(
        f"Loaded {len(contacts)} contacts."
    )

    if not contacts:

        print(
            "No contacts found in contacts.xlsx."
        )

        return

    # --------------------------------------------
    # Start browser
    # --------------------------------------------

    driver = start_browser()

    results = []

    try:

        # ----------------------------------------
        # Login
        # ----------------------------------------

        wait_for_login(driver)

        print()
        print(
            "Starting contact processing..."
        )

        # ----------------------------------------
        # Process contacts
        # ----------------------------------------

        for index, contact in enumerate(
            contacts,
            start=1
        ):

            print()
            print(
                f"[{index}/{len(contacts)}]"
            )

            result = process_contact(
                driver,
                contact
            )

            results.append(result)

            # Save after every contact
            # so results are not lost if the
            # program stops unexpectedly.
            save_json_report(results)
            save_excel_report(results)

            # Small delay between contacts
            time.sleep(2)

        print()
        print("=" * 60)
        print("PROCESS COMPLETED")
        print("=" * 60)

        print()
        print(
            f"JSON report: {JSON_REPORT}"
        )

        print(
            f"Excel report: {EXCEL_REPORT}"
        )

        print(
            f"Screenshots: {SCREENSHOT_DIR}"
        )

    except KeyboardInterrupt:

        print()
        print(
            "Bot stopped by user."
        )

        # Save whatever has already been processed
        save_json_report(results)
        save_excel_report(results)

    finally:

        print()
        print(
            "Browser will remain open for 5 seconds..."
        )

        time.sleep(5)

        driver.quit()


# ============================================================
# RUN PROGRAM
# ============================================================

if __name__ == "__main__":
    main()